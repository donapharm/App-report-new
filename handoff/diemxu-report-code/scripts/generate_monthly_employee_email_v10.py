from pathlib import Path
from openpyxl import load_workbook
import json, html
from collections import defaultdict
BASE=Path('/home/osboxes/.openclaw/workspace-report/t06_2026_v10_batch_18')
BASE.mkdir(parents=True, exist_ok=True)
T06=Path('/home/osboxes/.openclaw/workspace-main/webapp_donapharm/data/report_upload_data_20260601_20260630.json')
T05=Path('/home/osboxes/.openclaw/workspace-main/webapp_donapharm/data/report_upload_data_20260501_20260529.json')
XU=Path('/home/osboxes/.openclaw/media/inbound/Bao_cao_diem_xu_T06_Q2_2026_DEN_30_06_2026---39c5b2d0-7164-4500-aab6-8c1ae1108a9e.xlsx')
TARGETS_FILE=Path('/home/osboxes/.openclaw/workspace-report/t05_2026_integrated/batch_sale_t05/batch_targets_summary.json')
rows06=json.loads(T06.read_text(encoding='utf-8'))
rows05=json.loads(T05.read_text(encoding='utf-8'))
target_meta={t['code']:t for t in json.loads(TARGETS_FILE.read_text(encoding='utf-8'))}
allowed=['DN001','DN002','DN003','DN004','DN005','DN006','DN007','DN008','DN009','DN010','DN011','DN012','DN016','DN017','DN018','DN019','DN024','VP004']
# names fallback if needed
fallback={
'DN001':('Đặng Xuân Trung','Anh Trung'),'DN002':('Nguyễn Thị Hằng Nga','Chị Nga'),'DN003':('Nguyễn Trần Hoàng Anh','Anh Hoàng Anh'),'DN004':('Bùi Hoàng Ngọc Quyên','Chị Quyên'),'DN005':('Nguyễn Thị Dung','Chị Dung'),'DN006':('Nguyễn Trọng Hiếu','Anh Hiếu'),'DN007':('Trần Thị Kiều Linh','Chị Linh'),'DN008':('Đoàn Văn Triệu','Anh Triệu'),'DN009':('Trần Thị Thanh Huyền','Chị Huyền'),'DN010':('Trần Quốc Cường','Anh Cường'),'DN011':('Phan Tuấn','Anh Tuấn'),'DN012':('Đặng Thị Hồng Hạnh','Chị Hạnh'),'DN016':('Trần Thị Ngọc Ánh','Chị Ánh'),'DN017':('Trần Trịnh Kiều Oanh','Chị Oanh'),'DN018':('Nguyễn Huỳnh Phương Mai','Chị Mai'),'DN019':('Dương Thị Mến','Chị Mến'),'DN024':('Hoàng Văn Hà','Anh Hà'),'VP004':('Trần Hoàng Trung','Anh Trung')}
def val(r,*ks):
    for k in ks:
        if r.get(k) not in (None,''): return r.get(k)
    return ''
def rev(r): return float(r.get('REVENUE') or r.get('tong_tien') or 0)
def filt(rows,code): return [r for r in rows if str(val(r,'EMP_NUMBER','ma_nv')).upper()==code]
def money(n): return f"{round(float(n)):,}".replace(',', '.')+'đ'
def short_money(n):
    n=float(n); sign='-' if n<0 else ''; n=abs(n)
    if n>=1e9: return sign+f"{n/1e9:.2f}".replace('.', ',')+' tỷ'
    if n>=1e6: return sign+f"{n/1e6:.0f}".replace('.', ',')+' tr'
    return sign+money(n)
def pct(n): return f"{float(n):.1f}%".replace('.', ',')
def point(n): return f"{float(n or 0):.2f}".replace('.', ',')
def dp(cur,old): return None if not old else (cur-old)/old*100
def dtxt(cur,old):
    d=dp(cur,old)
    return '—' if d is None else ('+' if d>=0 else '')+pct(d)
def agg(rows,key):
    d=defaultdict(float)
    for r in rows: d[key(r)]+=rev(r)
    return sorted(d.items(), key=lambda x:x[1], reverse=True)
def cnt(rows,key): return len(set(key(r) for r in rows if key(r)))
def comp(cur,old,limit=6):
    cd,od=dict(cur),dict(old); out=[]
    for k in set(cd)|set(od):
        c,o=cd.get(k,0),od.get(k,0); diff=c-o
        if diff: out.append((k,diff,c,o,dp(c,o)))
    return sorted([x for x in out if x[1]>0], key=lambda x:x[1], reverse=True)[:limit], sorted([x for x in out if x[1]<0], key=lambda x:x[1])[:limit]
# xu records
wb=load_workbook(XU, data_only=True, read_only=True)
ws=wb['ChiTiet_Tung_NV']; headers=[c for c in next(ws.iter_rows(min_row=1,max_row=1,values_only=True))]; idx={h:i for i,h in enumerate(headers)}
xu_by_code={}
for row in ws.iter_rows(min_row=2, values_only=True):
    if row[idx['Mã NV']]: xu_by_code[row[idx['Mã NV']] ]={h:row[i] for h,i in idx.items()}
sw=wb['TongHop_App']; comp_total={}
for row in sw.iter_rows(values_only=True):
    if row and row[0]: comp_total[str(row[0])]=row
company_sur=float(comp_total['Thiếu/Dư toàn công ty'][2] or 0)
STYLE="""
body{font-family:Arial,Helvetica,sans-serif;color:#163032;background:#f2f6f6;margin:0}.wrap{max-width:880px;margin:10px auto;background:#fff;border:1px solid #d1e2e0;border-radius:14px;overflow:hidden}.topbrand{background:#fff;padding:12px 16px;border-bottom:1px solid #d7e8e6}.topbrandtbl{width:100%;border-collapse:collapse}.topbrandtbl td{border:0;padding:0;vertical-align:middle}.logotop{width:96px}.qrtop{width:64px;text-align:right}.logoTop{max-width:88px;height:auto;display:block}.qrTop{width:58px;height:58px;display:block;margin-left:auto}.brandmid{text-align:center;color:#00493f;font-size:16px;line-height:1.25}.brandmid span{font-size:12.5px;color:#476360}.head{background:#00493f;text-align:center;padding:17px 16px 18px;color:#fff}.eyebrow{font-size:12px;font-weight:800;letter-spacing:.7px;color:#b8fff1;text-transform:uppercase}.title{font-size:26px;font-weight:900;line-height:1.2;margin:5px 0;color:#fff}.sub{font-size:13.5px;line-height:1.35;color:#f4fffd;font-weight:600}.content{padding:16px 18px;line-height:1.45;font-size:14px}.note{background:#fff8eb;border-left:4px solid #d68b00;border-radius:8px;padding:8px 10px;margin:8px 0 14px;font-size:13px;color:#4a3900}.section{margin-top:18px}.section h3{font-size:16px;color:#005f52;border-bottom:2px solid #cfe4e2;padding-bottom:6px;margin:0 0 9px}.kpitbl{width:100%;border-collapse:separate;border-spacing:8px}.kpitbl td{border:0;padding:0;width:50%;vertical-align:top}.kpibox,.card{background:#f0faf8;border:1px solid #b7ddd8;border-radius:12px;padding:10px;min-height:64px}.cards{display:grid;grid-template-columns:repeat(2,1fr);gap:9px}.lab{font-size:12px;color:#315f5a;font-weight:800}.val{font-size:18px;font-weight:900;color:#00483f;margin-top:4px;line-height:1.2}.hint{font-size:12px;color:#4e6966;margin-top:3px}.analysis{background:#f7fbfb;border:1px solid #d7e8e6;border-radius:10px;padding:10px 11px;margin:8px 0}.analysis b{color:#00493f}table{border-collapse:collapse;width:100%;font-size:13px}td,th{border:1px solid #d7e8e6;padding:6px 7px;vertical-align:top}th{background:#e9f6f4;color:#00483f;text-align:left}.r{text-align:right;white-space:nowrap}.neg{color:#b00020;font-weight:800}.pos{color:#087a42;font-weight:800}.pill{display:inline-block;border-radius:99px;padding:3px 8px;background:#eaf7f4;color:#00493f;font-weight:800}.focus{background:#f0faf8;border-left:4px solid #087565;border-radius:8px;padding:9px 12px;margin:8px 0}.foot{margin-top:14px;padding:10px 12px;background:#f4f8f8;border-top:1px solid #d7e8e6;color:#476360;font-size:12.5px;line-height:1.45}@media(max-width:720px){.wrap{margin:0;border-radius:0}.topbrand{padding:10px 12px}.logotop{width:78px}.qrtop{width:52px}.logoTop{max-width:72px}.qrTop{width:46px;height:46px}.brandmid{font-size:14px}.brandmid span{font-size:11px}.head{padding:15px 12px 16px}.title{font-size:24px}.content{padding:13px 10px;font-size:13.5px}.kpitbl{border-spacing:6px}.kpibox{padding:8px;min-height:58px}.cards{grid-template-columns:repeat(2,1fr);gap:7px}.card{padding:8px}.val{font-size:15.5px}table{font-size:12px}td,th{padding:5px}.r{white-space:normal}}
"""
def top_table(title,pairs,total,limit=8):
    s=f"<div class='section'><h3>{title}</h3><table><tr><th>#</th><th>Nội dung</th><th class='r'>Doanh thu</th><th class='r'>%</th></tr>"
    for i,(k,v) in enumerate(pairs[:limit],1): s+=f"<tr><td>{i}</td><td>{html.escape(str(k))}</td><td class='r'>{money(v)}</td><td class='r'>{pct(v/total*100 if total else 0)}</td></tr>"
    return s+'</table></div>'
def dec_table(title,rows,limit=5):
    s=f"<div class='section'><h3>{title}</h3><table><tr><th>Nội dung</th><th class='r'>T05</th><th class='r'>T06</th><th class='r'>Giảm</th></tr>"
    if not rows:
        s += "<tr><td colspan='4'>Không có giảm đáng kể.</td></tr>"
    for k,d,c,o,p in rows[:limit]: s+=f"<tr><td>{html.escape(str(k))}</td><td class='r'>{short_money(o)}</td><td class='r'>{short_money(c)}</td><td class='r neg'>{short_money(abs(d))}</td></tr>"
    return s+'</table></div>'
summary=[]
for code in allowed:
    meta=target_meta.get(code,{})
    name=meta.get('name') or fallback.get(code,(code,'Anh/Chị'))[0]
    salute=meta.get('salute') or fallback.get(code,(name,'Anh/Chị'))[1]
    email=meta.get('email',''); phone=meta.get('phone','')
    r06=filt(rows06,code); r05=filt(rows05,code)
    total06=sum(map(rev,r06)); total05=sum(map(rev,r05))
    routes06=agg(r06, lambda r: val(r,'TUYEN','tuyen') or '#N/A'); routes05=dict(agg(r05, lambda r: val(r,'TUYEN','tuyen') or '#N/A'))
    units06=agg(r06, lambda r: val(r,'DONVI','ten_vt','donvi') or '(trống)'); units05=agg(r05, lambda r: val(r,'DONVI','ten_vt','donvi') or '(trống)')
    items06=agg(r06, lambda r: val(r,'ITEM_NAME','IIT_NAME','NAME','ten_hang') or '(trống)'); items05=agg(r05, lambda r: val(r,'ITEM_NAME','IIT_NAME','NAME','ten_hang') or '(trống)')
    vendors06=agg(r06, lambda r: val(r,'NHA_THAU','VEN_NAME','ten_nha_thau') or '(trống)')
    unit_inc,unit_dec=comp(units06,units05,6); item_inc,item_dec=comp(items06,items05,6)
    route="<div class='section'><h3>2. Cơ cấu tuyến</h3><table><tr><th>Tuyến</th><th class='r'>T06</th><th class='r'>So T05</th><th class='r'>Tỷ trọng</th></tr>"
    for k,v in routes06:
        o=routes05.get(k,0); d=v-o
        route+=f"<tr><td>{html.escape(str(k))}</td><td class='r'>{money(v)}</td><td class='r {'pos' if d>=0 else 'neg'}'>{('+' if d>=0 else '')}{short_money(d)}</td><td class='r'>{pct(v/total06*100 if total06 else 0)}</td></tr>"
    route+='</table></div>'
    rec=xu_by_code.get(code)
    if rec:
        m_point=float(rec['Điểm tháng'] or 0); m_xu=float(rec['Xu tháng'] or 0); q_point=float(rec['Điểm quý'] or 0); q_xu=float(rec['Xu tổng quý'] or 0); q_diff=q_xu-q_point; m_diff=m_xu-m_point
        xu_sec=f"""<div class='section'><h3>7. Điểm doanh thu & xu chi tiêu</h3><div class='cards'><div class='card'><div class='lab'>Điểm T06</div><div class='val'>{point(m_point)}</div><div class='hint'>Xu T06: {point(m_xu)} • lệch <span class='{'pos' if m_diff>=0 else 'neg'}'>{point(m_diff)}</span></div></div><div class='card'><div class='lab'>Điểm Q2</div><div class='val'>{point(q_point)}</div><div class='hint'>Xu tổng Q2: {point(q_xu)}</div></div><div class='card'><div class='lab'>Trạng thái cá nhân</div><div class='val'><span class='pill'>{html.escape(str(rec['Trạng thái cá nhân 100%']))}</span></div><div class='hint'>Chênh Q2: <span class='{'pos' if q_diff>=0 else 'neg'}'>{('+' if q_diff>=0 else '')}{point(q_diff)} xu</span></div></div><div class='card'><div class='lab'>Truy thu cá nhân</div><div class='val'>{money(rec['Truy thu nếu xét cá nhân <90%'])}</div><div class='hint'>Tỷ lệ Q2: {point(rec['Tỷ lệ quý %'])}%</div></div></div><div class='analysis'><b>Lưu ý:</b> Toàn công ty Q2 đang dư <b>{point(company_sur)} xu</b>, không kết luận phạt theo tổng từ thiếu/dư cá nhân.</div></div>"""
    else:
        xu_sec="<div class='section'><h3>7. Điểm doanh thu & xu chi tiêu</h3><div class='analysis'>Chưa có dữ liệu điểm/xu cho mã này trong file Q2.</div></div>"
    risks_units=', '.join(x[0] for x in unit_dec[:3]) or 'không có điểm giảm lớn'
    risks_items=', '.join(x[0] for x in item_dec[:3]) or 'không có mã giảm lớn'
    summary_sec=f"""<div class='section'><h3>8. Nhận xét nhanh & hành động T07</h3><div class='analysis'><b>Kết quả:</b> {code} đạt <b>{money(total06)}</b>, <b>{dtxt(total06,total05)}</b> so với T05.</div><div class='analysis'><b>Rủi ro cần xử lý:</b> đơn vị giảm mạnh: {html.escape(risks_units)}; mặt hàng giảm sâu: {html.escape(risks_items)}.</div><p class='focus'><b>Việc cần làm T07:</b> gọi lại nhóm đơn vị giảm trong 7 ngày đầu; kiểm tra tồn/nhu cầu các mã giảm sâu; giữ nhịp các điểm doanh thu lớn trong tháng.</p></div>"""
    html_body=f"""<!doctype html><html><head><meta charset='utf-8'><style>{STYLE}</style></head><body><div class='wrap'><div class='topbrand'><table class='topbrandtbl' role='presentation'><tr><td class='logotop'><img class='logoTop' src='cid:logo_dona' alt='DONAPHARM'></td><td class='brandmid'><b>DONAPHARM</b><br><span>Văn phòng CEO • App Report nội bộ</span></td><td class='qrtop'><img class='qrTop' src='cid:qr_zalo' alt='Zalo OA'></td></tr></table></div><div class='head'><div class='eyebrow'>BÁO CÁO DOANH THU</div><div class='title'>Tháng 06/2026</div><div class='sub'>{html.escape(name)} – {code}<br>01/06–30/06/2026 • So sánh T05</div></div><div class='content'><p>Kính gửi <b>{html.escape(salute)} ({code})</b>,</p><p>Văn phòng CEO gửi báo cáo doanh thu tháng 06/2026 đã lọc riêng theo tài khoản/NV <b>{code}</b>.</p><p class='note'><b>Nguồn dữ liệu:</b> App Report 06.2026 - Tuần 27 ✅ + file Điểm/Xu Q2 đến 30/06/2026. Báo cáo không chứa chi phí, lợi nhuận, margin.</p><div class='section'><h3>1. Dashboard {code}</h3><table class='kpitbl' role='presentation'><tr><td><div class='kpibox'><div class='lab'>Doanh thu T06</div><div class='val'>{short_money(total06)}</div><div class='hint'>{money(total06)}</div></div></td><td><div class='kpibox'><div class='lab'>So với T05</div><div class='val {'pos' if total06-total05>=0 else 'neg'}'>{dtxt(total06,total05)}</div><div class='hint'>T05: {short_money(total05)}</div></div></td></tr><tr><td><div class='kpibox'><div class='lab'>Số dòng</div><div class='val'>{len(r06)}</div><div class='hint'>T05: {len(r05)}</div></div></td><td><div class='kpibox'><div class='lab'>Đơn vị / Mặt hàng</div><div class='val'>{cnt(r06,lambda r: val(r,'DONVI','ten_vt','donvi'))} / {cnt(r06,lambda r: val(r,'ITEM_NAME','IIT_NAME','NAME','ten_hang'))}</div><div class='hint'>Nhà thầu: {cnt(r06,lambda r: val(r,'NHA_THAU','VEN_NAME','ten_nha_thau'))}</div></div></td></tr></table></div>{route}{top_table('3. Top đơn vị T06',units06,total06,8)}{dec_table('4. Đơn vị giảm mạnh so với T05',unit_dec,5)}{dec_table('5. Mặt hàng giảm sâu so với T05',item_dec,5)}{top_table('6. Top mặt hàng T06',items06,total06,8)}{xu_sec}{summary_sec}<p style='margin-top:22px'>Trân trọng,<br><b>ĐẶNG XUÂN TRUNG</b><br>CEO — Công ty TNHH Dược phẩm DONAPHARM<br>Hotline: 0886.396.668</p><div class='foot'><b>E-Mail này được gửi từ Văn phòng CEO DONAPHARM</b> thông qua hệ thống App Report nội bộ. Nội dung chỉ dành cho người nhận, vui lòng không chuyển tiếp ra bên ngoài khi chưa được phép.</div></div></div></body></html>"""
    plain=f"DONAPHARM - Báo cáo doanh thu T06/2026\nKính gửi {salute} ({code})\nDoanh thu T06: {money(total06)} ({dtxt(total06,total05)} so T05)\nĐiểm/Xu: " + (f"Điểm Q2 {point(rec['Điểm quý'])}, Xu Q2 {point(rec['Xu tổng quý'])}, trạng thái {rec['Trạng thái cá nhân 100%']}" if rec else 'chưa có') + "\nE-Mail này được gửi từ Văn phòng CEO DONAPHARM.\n"
    (BASE/f'{code}_T06_2026_V10.html').write_text(html_body,encoding='utf-8')
    (BASE/f'{code}_T06_2026_V10.txt').write_text(plain,encoding='utf-8')
    summary.append({'code':code,'name':name,'salute':salute,'email':email,'phone':phone,'revenue_t06':total06,'revenue_t05':total05,'delta_pct':dp(total06,total05),'has_xu':bool(rec),'html':str(BASE/f'{code}_T06_2026_V10.html'),'txt':str(BASE/f'{code}_T06_2026_V10.txt')})
json.dump({'targets':summary,'allowedCodes':allowed,'generatedAt':'2026-06-30T21:25:00+07:00','template':'V10'}, open(BASE/'batch_targets_v10_18.json','w',encoding='utf-8'), ensure_ascii=False, indent=2)
# preview markdown
md=['# Batch gửi báo cáo doanh thu T06/2026 - V10','',f'- Số người gửi: {len(summary)}','- Chỉ gửi đúng danh sách CEO yêu cầu: '+', '.join(allowed),'','| Mã | Tên | Email | Phone | DT T06 | So T05 | Có điểm/xu |','|---|---|---|---|---:|---:|---|']
for t in summary:
    d='—' if t['delta_pct'] is None else ('+' if t['delta_pct']>=0 else '')+pct(t['delta_pct'])
    md.append(f"| {t['code']} | {t['name']} | {t['email']} | {t['phone']} | {money(t['revenue_t06'])} | {d} | {'Có' if t['has_xu'] else 'Không'} |")
(BASE/'PREVIEW_BATCH_V10_18.md').write_text('\n'.join(md)+'\n',encoding='utf-8')
print((BASE/'PREVIEW_BATCH_V10_18.md').read_text(encoding='utf-8'))
