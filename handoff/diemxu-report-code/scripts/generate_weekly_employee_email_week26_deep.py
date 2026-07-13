#!/usr/bin/env python3
from openpyxl import load_workbook
from pathlib import Path
from collections import defaultdict, Counter
from datetime import datetime, date
import html, json, math

WS=Path('/home/osboxes/.openclaw/workspace-office')
SALES_JUN=Path('/home/osboxes/.openclaw/media/inbound/BAO_CAO_DOANH_SO_TUAN_26_22.06.26-26.06.26---72d72f3e-f867-4494-b44f-2a582c796040.xlsx')
SALES_MAY=Path('/home/osboxes/.openclaw/media/inbound/BAO_CAO_DOANH_SO_TUAN_22_01.05.26-29.05.26---8c444c17-6174-4d32-b178-c9468339b5fe.xlsx')
POINTS=Path('/home/osboxes/.openclaw/media/inbound/Bao_cao_diem_xu_T06_Q2_2026_DA_DOI_CHIEU_ANH_APP_2238fe70_5f---de35ff33-e422-411c-a5a8-e2b5cf0ae581.xlsx')
OUT=WS/'reports/week26_sales_reports_20260626_deep'
HTML_DIR=OUT/'html'; TXT_DIR=OUT/'txt'
HTML_DIR.mkdir(parents=True,exist_ok=True); TXT_DIR.mkdir(parents=True,exist_ok=True)

EMP_NAMES={
'DN001':'Đặng Xuân Trung','DN002':'Nguyễn Thị Hằng Nga','DN003':'Nguyễn Trần Hoàng Anh','DN004':'Bùi Hoàng Ngọc Quyên','DN005':'Nguyễn Thị Dung','DN006':'Nguyễn Trọng Hiếu','DN007':'Trần Thị Kiều Linh','DN008':'Đoàn Văn Triệu','DN009':'Trần Thị Thanh Huyền','DN010':'Trần Quốc Cường','DN011':'Phan Tuấn','DN012':'Đặng Thị Hồng Hạnh','DN016':'Trần Thị Ngọc Ánh','DN017':'Trần Trịnh Kiều Oanh','DN018':'Nguyễn Huỳnh Phương Mai','DN019':'Dương Thị Mến','DN021':'Nguyễn Thị Yến Nhi','DN022':'Đồng Văn Toản','DN023':'Nguyễn Đức Tuấn','DN024':'Hoàng Văn Hà','VP004':'VP004'
}

def esc(x): return html.escape(str(x if x is not None else ''))
def money(n):
    try: n=float(n or 0)
    except Exception: n=0
    return f"{round(n):,}".replace(',', '.')+'đ'
def num(n,dec=2):
    try: n=float(n or 0)
    except Exception: n=0
    s=f"{n:,.{dec}f}".replace(',', 'X').replace('.', ',').replace('X','.')
    return s
def pct(n): return num(n,1)+'%'
def dt(v):
    if isinstance(v,datetime): return v.date()
    if isinstance(v,date): return v
    return None

def load_sales(path, start_month, end_day):
    wb=load_workbook(path, read_only=True, data_only=True)
    ws=wb[wb.sheetnames[0]]
    rows=[]
    for r in ws.iter_rows(min_row=4, values_only=True):
        if not r or not r[0] or not r[3]: continue
        d=dt(r[0])
        if not d or d.month!=start_month or d.day>end_day: continue
        rows.append({
            'date':d,'tuyen':str(r[1] or '').strip() or 'Khác','unit':str(r[2] or '').strip(),
            'nv':str(r[3] or '').strip(),'code':str(r[4] or '').strip(),'product':str(r[5] or '').strip(),
            'vendor':str(r[6] or '').strip(),'uom':str(r[7] or '').strip(),'qty':float(r[8] or 0),
            'price':float(r[9] or 0),'rev':float(r[10] or 0)
        })
    return rows

def load_points(path):
    wb=load_workbook(path, read_only=True, data_only=True)
    ws=wb['ChiTiet_Tung_NV']
    headers=[c.value for c in next(ws.iter_rows(min_row=1,max_row=1))]
    out={}
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not any(r): continue
        d=dict(zip(headers,r)); code=str(d.get('Mã NV') or '').strip()
        if not code: continue
        # VP009 duplicated in source: keep first; manifest will note source duplicate.
        out.setdefault(code,d)
        EMP_NAMES[code]=str(d.get('Tên NV') or EMP_NAMES.get(code,code))
    return out

jun=load_sales(SALES_JUN,6,26)
may=load_sales(SALES_MAY,5,26)
points=load_points(POINTS)
all_codes=sorted(set([r['nv'] for r in jun]) | set(points.keys()))

# Remove pure VP/no-sales/no-KPI duplicates? Keep all with point file for complete package.

def aggregate(rows, code):
    rr=[r for r in rows if r['nv']==code]
    total=sum(r['rev'] for r in rr)
    return {
        'rows':rr,'total':total,'line_count':len(rr),
        'units':len({r['unit'] for r in rr if r['unit']}),
        'products':len({r['product'] for r in rr if r['product']}),
        'vendors':len({r['vendor'] for r in rr if r['vendor']}),
        'by_tuyen':sum_by(rr,'tuyen'),'by_unit':sum_by(rr,'unit'),'by_product':sum_by(rr,'product'),
        'by_day':sum_by(rr, lambda r:r['date'].strftime('%d/%m'))
    }

def sum_by(rows, key):
    d=defaultdict(float)
    for r in rows:
        k=key(r) if callable(key) else r[key]
        d[k]+=r['rev']
    return dict(d)

def top_items(d, n=8, reverse=True):
    return sorted(d.items(), key=lambda x:x[1], reverse=reverse)[:n]

def compare_top(cur, prev, n=6, positive=True):
    keys=set(cur)|set(prev)
    arr=[]
    for k in keys:
        c=cur.get(k,0); p=prev.get(k,0); delta=c-p
        if positive and delta<=0: continue
        if not positive and delta>=0: continue
        arr.append((k,c,p,delta))
    arr.sort(key=lambda x:x[3], reverse=positive)
    return arr[:n]

def list_sentence(items, kind):
    if not items: return f'Chưa có {kind} nổi bật.'
    out=[]
    for k,c,p,d in items[:5]:
        if c==0 and p>0: out.append(f'{k} chưa phát sinh lại')
        else: out.append(f'{k} {"tăng" if d>0 else "giảm"} {money(abs(d))}')
    return '; '.join(out)

def report_html(code):
    name=EMP_NAMES.get(code,code)
    a=aggregate(jun,code); b=aggregate(may,code)
    total=a['total']; prev=b['total']; delta=total-prev; rate=(delta/prev*100) if prev else (100 if total else 0)
    p=points.get(code,{})
    diem_t=float(p.get('Điểm tháng') or 0); xu_t=float(p.get('Xu tháng') or 0)
    diem_q=float(p.get('Điểm quý') or 0); xu_q=float(p.get('Xu quý') or 0); xu_du=float(p.get('Xu dư quý trước') or 0); xu_tong=float(p.get('Xu tổng quý') or (xu_q+xu_du))
    thieu=float(p.get('Thiếu xu') or max(0,diem_q-xu_tong)); du=float(p.get('Dư xu') or max(0,xu_tong-diem_q))
    tyle_q=float(p.get('Tỷ lệ quý %') or ((xu_tong/diem_q*100) if diem_q else 0))
    warn=str(p.get('Cảnh báo cá nhân <90%') or 'KHÔNG')
    need90=float(p.get('Thiếu để đạt 90%') or 0); truy=float(p.get('Truy thu nếu xét cá nhân <90%') or 0)
    top_units=top_items(a['by_unit'],8); top_prods=top_items(a['by_product'],8)
    up_units=compare_top(a['by_unit'], b['by_unit'],5,True); down_units=compare_top(a['by_unit'], b['by_unit'],5,False)
    up_prod=compare_top(a['by_product'], b['by_product'],5,True); down_prod=compare_top(a['by_product'], b['by_product'],5,False)
    maxday=max(a['by_day'].values() or [1])
    tuyen_keys=sorted(set(a['by_tuyen'])|set(b['by_tuyen']))
    colors=['#087565','#0d9488','#f59e0b','#64748b']
    def rows_top(items,total):
        return ''.join(f"<tr><td>{i}</td><td>{esc(k)}</td><td class='r'>{money(v)}</td><td class='r'>{pct(v/total*100 if total else 0)}</td></tr>" for i,(k,v) in enumerate(items,1)) or "<tr><td colspan='4'>Chưa có phát sinh.</td></tr>"
    def rows_cmp(items):
        return ''.join(f"<tr><td>{esc(k)}</td><td class='r'>{money(c)}</td><td class='r'>{money(pv)}</td><td class='r {'pos' if d>0 else 'neg'}'>{money(d)}</td><td>{'Tăng' if d>0 else ('Chưa phát sinh lại' if c==0 else 'Giảm')}</td></tr>" for k,c,pv,d in items) or "<tr><td colspan='5'>Chưa có biến động nổi bật.</td></tr>"
    tuyen_bars=''
    for idx,k in enumerate(tuyen_keys):
        c=a['by_tuyen'].get(k,0); pv=b['by_tuyen'].get(k,0); share=(c/total*100 if total else 0)
        if c==0 and pv==0: continue
        tuyen_bars += f"<div class='bar-row'><div class='bar-label'>{esc(k)}</div><div class='bar-wrap'><div class='bar' style='width:{min(100,share):.1f}%;background:{colors[idx%len(colors)]}'></div></div><div class='bar-val'>{money(c)}<br><span>{pct(share)}</span></div><div class='bar-sub'>T05: {money(pv)}</div></div>"
    day_chart=''.join(f"<div class='day'><div class='col' style='height:{max(3,round(v/maxday*96))}px'></div><div class='day-val'>{money(v)}</div><div class='day-lab'>{esc(k)}</div></div>" for k,v in sorted(a['by_day'].items(), key=lambda x: datetime.strptime(x[0],'%d/%m')))
    top_unit_names=[k for k,v in top_units[:3]]
    focus_down=list_sentence(down_units,'đơn vị giảm')
    focus_up=list_sentence(up_prod,'mã tăng')
    focus_down_prod=list_sentence(down_prod,'mã giảm')
    exists=[]
    if delta<0: exists.append(f"Doanh thu giảm {money(abs(delta))} so cùng kỳ T05 ({pct(rate)}).")
    if warn.upper()=='CÓ': exists.append(f"Cảnh báo KPI cá nhân dưới 90%: tỷ lệ quý {pct(tyle_q)}, thiếu để đạt 90% là {num(need90)} xu; truy thu tham chiếu {money(truy)} nếu xét cá nhân.")
    elif thieu>0: exists.append(f"Điểm/xu quý còn thiếu {num(thieu)} xu; cần tăng chi tiêu hợp lệ hoặc bù bằng doanh thu/xu trong các kỳ tới.")
    else: exists.append(f"Điểm/xu quý đang dư {num(du)} xu, cần giữ nhịp để không tụt cuối kỳ.")
    if down_units: exists.append('Đơn vị giảm/chưa phát sinh lại: '+focus_down+'.')
    if down_prod: exists.append('Mã hàng giảm/chưa phát sinh lại: '+focus_down_prod+'.')
    exists_html=''.join('<li>'+esc(x)+'</li>' for x in exists)
    klass='pos' if delta>=0 else 'neg'
    point_note='Dữ liệu điểm/xu lấy từ file “Báo cáo điểm/xu T06-Q2 đã đối chiếu ảnh app”. File hiện không có cột chi tiêu hợp lệ/bill chi tiết, nên email chỉ hiển thị điểm/xu/thiếu-dư/tỷ lệ.'
    return f"""<!doctype html><html><head><meta charset='utf-8'><style>
body{{margin:0;background:#f6fbfb;font-family:Arial,Helvetica,sans-serif;color:#163235}}.wrap{{max-width:960px;margin:0 auto;background:#fff;border-radius:18px;overflow:hidden;border:1px solid #d8eeee;box-shadow:0 8px 28px rgba(0,80,90,.08)}}.banner{{background:linear-gradient(135deg,#e8fbf8 0%,#c9f2ec 48%,#74d0c4 100%);padding:16px 24px 18px}}.brand{{font-size:28px;font-weight:900;letter-spacing:.6px;color:#087565}}.title{{font-size:23px;font-weight:900;color:#005f52;margin-top:5px}}.period{{font-size:13px;color:#245f5a;margin-top:4px}}.content{{padding:28px 32px 30px;line-height:1.62;font-size:15.2px}}.note{{background:#fff8e8;border-left:4px solid #d99a00;border-radius:10px;padding:12px 16px}}.section{{margin-top:25px}}.section h3{{font-size:18px;color:#087565;margin:0 0 12px;padding-bottom:7px;border-bottom:2px solid #d8eeee}}.overview-table{{width:100%;border-collapse:separate;border-spacing:10px;table-layout:fixed;margin-left:-10px;margin-right:-10px}}.overview-table td{{border:0;padding:0;width:50%;vertical-align:top}}.kpi-card{{background:linear-gradient(180deg,#eefbf8 0%,#f9fffe 100%);border:1px solid #cceee8;border-radius:14px;padding:15px 12px;text-align:center;min-height:86px;box-shadow:0 2px 8px rgba(0,95,82,.05)}}.kpi-card.primary{{background:linear-gradient(180deg,#e3f8f4 0%,#f7fffd 100%);border-color:#9fe1d8}}.kpi-card .label{{font-size:12px;color:#456;text-transform:uppercase;letter-spacing:.25px}}.kpi-card .value{{font-size:20px;font-weight:900;color:#005f52;margin-top:6px;line-height:1.2}}.kpi-card .delta{{font-size:12.5px;margin-top:5px}}.pos{{color:#087565;font-weight:800}}.neg{{color:#b42318;font-weight:800}}table{{border-collapse:collapse;width:100%;font-size:13.5px}}th{{background:#e7f7f4;color:#005f52;text-align:left}}td,th{{border:1px solid #d8eeee;padding:8px 9px;vertical-align:top}}.r{{text-align:right}}.highlight{{background:#eefbf8;border-left:4px solid #087565;border-radius:10px;padding:12px 16px}}.warn{{background:#fff5f0;border-left:4px solid #ef7b45;border-radius:10px;padding:12px 16px}}.small{{font-size:12.3px;color:#667}}.bar-row{{display:grid;grid-template-columns:145px 1fr 145px 120px;gap:10px;align-items:center;margin:9px 0}}.bar-label{{font-weight:700}}.bar-wrap{{height:18px;background:#e8f4f2;border-radius:9px;overflow:hidden}}.bar{{height:18px;border-radius:9px}}.bar-val{{text-align:right;font-weight:800;color:#005f52}}.bar-val span,.bar-sub{{font-size:12px;color:#667}}.day-chart{{display:flex;align-items:flex-end;gap:8px;height:132px;padding:12px 8px 2px;border:1px solid #d8eeee;border-radius:12px;background:#fbfffe;overflow-x:auto}}.day{{text-align:center;min-width:54px}}.col{{width:28px;background:#0d9488;border-radius:6px 6px 0 0;margin:0 auto}}.day-val{{font-size:10px;color:#245f5a;margin-top:4px;white-space:nowrap}}.day-lab{{font-size:11px;color:#667}}.grid2{{display:grid;grid-template-columns:1fr 1fr;gap:14px}}.todo li{{margin:6px 0}}@media(max-width:760px){{.content{{padding:22px 16px}}.cards{{grid-template-columns:repeat(2,1fr)}}.grid2{{grid-template-columns:1fr}}.brand{{font-size:23px}}.title{{font-size:18px}}table{{font-size:12.5px}}.banner{{padding:14px 14px}}.bar-row{{grid-template-columns:92px 1fr 92px}}.bar-sub{{display:none}}}}
</style></head><body><div style='padding:24px 0'><div class='wrap'><div class='banner'><table style='border:0;width:100%'><tr><td style='border:0;width:82px;vertical-align:top'><img src='cid:logo_dona' alt='DONAPHARM' style='width:64px;max-width:64px;height:auto;display:block;border:0;background:transparent'></td><td style='border:0;vertical-align:middle'><div class='brand'>DONAPHARM</div><div class='title'>Báo cáo doanh thu Tuần 26</div><div class='period'>{esc(name)} – {esc(code)} | Kỳ 01/06–26/06/2026 | So sánh cùng kỳ T05</div></td><td style='border:0;width:58px;text-align:right;vertical-align:top'><div style='width:58px;text-align:center;line-height:1.15'><img src='cid:qr_zalo' alt='QR Zalo' style='width:46px;height:46px;display:block;margin:0 auto;border:0;background:transparent'><div style='font-size:7.5px;color:#245f5a;margin-top:3px;white-space:nowrap;text-align:center;width:58px;max-width:58px;overflow:hidden'>Quét QR Zalo OA</div></div></td></tr></table></div><div class='content'><p>Kính gửi <b>Anh/Chị {esc(name)} ({esc(code)})</b>,</p><p>Văn phòng CEO gửi báo cáo doanh thu Tuần 26, lọc riêng theo mã nhân viên <b>{esc(code)}</b>. Báo cáo bám mẫu tháng/tuần chuẩn: doanh thu, tuyến, điểm/xu, đơn vị, mặt hàng, nhà thầu, biến động cùng kỳ và kiến nghị hành động.</p><p class='note'><b>Nguồn dữ liệu:</b> App Report/Excel doanh số <b>01/06–26/06/2026</b> + Excel điểm/xu T06-Q2 đã đối chiếu ảnh app. Báo cáo không chứa chi phí, giá vốn, lợi nhuận, margin hoặc dữ liệu nhạy cảm.</p>
<div class='section'><h3>1. Tổng quan kết quả</h3><table class='overview-table' role='presentation' cellpadding='0' cellspacing='0'><tr><td><div class='kpi-card primary'><div class='label'>Doanh thu</div><div class='value'>{money(total)}</div><div class='delta {klass}'>{pct(rate)} so T05</div></div></td><td><div class='kpi-card'><div class='label'>Chênh lệch so T05</div><div class='value {klass}'>{money(delta)}</div><div class='delta'>T05: {money(prev)}</div></div></td></tr><tr><td><div class='kpi-card'><div class='label'>Số dòng</div><div class='value'>{a['line_count']}</div><div class='delta'>T05: {b['line_count']}</div></div></td><td><div class='kpi-card'><div class='label'>Đơn vị</div><div class='value'>{a['units']}</div><div class='delta'>T05: {b['units']}</div></div></td></tr><tr><td><div class='kpi-card'><div class='label'>Mặt hàng</div><div class='value'>{a['products']}</div><div class='delta'>T05: {b['products']}</div></div></td><td><div class='kpi-card'><div class='label'>Nhà thầu</div><div class='value'>{a['vendors']}</div><div class='delta'>T05: {b['vendors']}</div></div></td></tr></table></div>
<div class='section'><h3>2. Điểm doanh thu & xu chi tiêu</h3><p class='highlight'>{esc(point_note)}</p><table><tr><th>Kỳ</th><th class='r'>Doanh thu</th><th class='r'>Điểm DT</th><th class='r'>Xu kỳ này</th><th class='r'>Xu Q2</th><th class='r'>Xu dư quý trước</th><th class='r'>Xu tổng quý</th><th class='r'>Thiếu xu</th><th class='r'>Dư xu</th><th class='r'>Hoàn thành quý</th></tr><tr><td>Tháng 06 đến 26/06</td><td class='r'>{money(total)}</td><td class='r'><b>{num(diem_t)}</b></td><td class='r'>{num(xu_t)}</td><td class='r'>{num(xu_q)}</td><td class='r'>{num(xu_du)}</td><td class='r'>{num(xu_tong)}</td><td class='r neg'>{num(thieu)}</td><td class='r pos'>{num(du)}</td><td class='r'>{pct(tyle_q)}</td></tr></table>{('<p class="warn"><b>Cảnh báo cá nhân &lt;90%:</b> còn thiếu '+num(need90)+' xu để đạt 90%; truy thu tham chiếu '+money(truy)+' nếu xét theo từng cá nhân.</p>') if warn.upper()=='CÓ' else ''}</div>
<div class='section'><h3>3. Phân tích tuyến CL / NCL / NT</h3><p class='highlight'>Cơ cấu tuyến giúp nhận diện NV đang phụ thuộc vào CL/NCL/NT nào và tuyến nào cần kéo lại trong tuần tới.</p>{tuyen_bars}<table><tr><th>Tuyến</th><th class='r'>T06 đến 26/06</th><th class='r'>Tỷ trọng</th><th class='r'>Cùng kỳ T05</th><th class='r'>Chênh lệch</th></tr>{''.join(f"<tr><td>{esc(k)}</td><td class='r'>{money(a['by_tuyen'].get(k,0))}</td><td class='r'>{pct((a['by_tuyen'].get(k,0)/total*100) if total else 0)}</td><td class='r'>{money(b['by_tuyen'].get(k,0))}</td><td class='r {'pos' if (a['by_tuyen'].get(k,0)-b['by_tuyen'].get(k,0))>=0 else 'neg'}'>{money(a['by_tuyen'].get(k,0)-b['by_tuyen'].get(k,0))}</td></tr>" for k in tuyen_keys)}</table></div>
<div class='section'><h3>4. Biểu đồ doanh thu theo ngày</h3><div class='day-chart'>{day_chart or '<div>Chưa có phát sinh doanh thu trong kỳ.</div>'}</div></div>
<div class='section'><h3>5. Top đơn vị và top mặt hàng</h3><div class='grid2'><div><h4>Top đơn vị</h4><table><tr><th>Hạng</th><th>Đơn vị</th><th class='r'>DT</th><th class='r'>Tỷ trọng</th></tr>{rows_top(top_units,total)}</table></div><div><h4>Top mặt hàng</h4><table><tr><th>Hạng</th><th>Mặt hàng</th><th class='r'>DT</th><th class='r'>Tỷ trọng</th></tr>{rows_top(top_prods,total)}</table></div></div></div>
<div class='section'><h3>6. So sánh tăng/giảm so cùng kỳ T05</h3><div class='grid2'><div><h4>Đơn vị tăng mạnh</h4><table><tr><th>Đơn vị</th><th class='r'>T06</th><th class='r'>T05</th><th class='r'>Chênh</th><th>Ghi chú</th></tr>{rows_cmp(up_units)}</table></div><div><h4>Đơn vị giảm/chưa phát sinh</h4><table><tr><th>Đơn vị</th><th class='r'>T06</th><th class='r'>T05</th><th class='r'>Chênh</th><th>Ghi chú</th></tr>{rows_cmp(down_units)}</table></div></div></div>
<div class='section'><h3>7. Tồn tại cần xử lý</h3><ul class='todo'>{exists_html}</ul></div>
<div class='section'><h3>8. Kiến nghị hành động tuần tới</h3><table><tr><th>Nhóm việc</th><th>Khuyến nghị cụ thể</th><th>Thời hạn</th></tr><tr><td>Giữ điểm lớn</td><td>{esc(', '.join(top_unit_names) or 'Chưa có điểm lớn rõ ràng')}</td><td>48 giờ</td></tr><tr><td>Kéo lại điểm giảm</td><td>{esc(focus_down)}</td><td>Trước giữa tuần</td></tr><tr><td>Đẩy mã có sức kéo</td><td>{esc(list_sentence(up_prod,'mã tăng'))}</td><td>3 ngày đầu tuần</td></tr><tr><td>Chặn tụt mã cũ</td><td>{esc(focus_down_prod)}</td><td>Trong tuần</td></tr></table></div>
<p style='margin-top:28px'>Trân trọng,<br><b>ĐẶNG XUÂN TRUNG</b><br>CEO — Công ty TNHH Dược phẩm DONAPHARM<br>Hotline: 0886.396.668</p><p class='small'><em>E-mail này được gửi từ văn phòng CEO DONAPHARM với sự hỗ trợ AI agent Donapharm. Đây là bản test DN001 để CEO góp ý trước khi gửi chính thức cho nhân viên.</em></p></div></div></div></body></html>"""

def report_txt(code):
    name=EMP_NAMES.get(code,code); a=aggregate(jun,code); b=aggregate(may,code); p=points.get(code,{})
    total=a['total']; prev=b['total']; delta=total-prev; rate=(delta/prev*100) if prev else (100 if total else 0)
    diem_t=float(p.get('Điểm tháng') or 0); xu_t=float(p.get('Xu tháng') or 0); diem_q=float(p.get('Điểm quý') or 0); xu_q=float(p.get('Xu quý') or 0); xu_du=float(p.get('Xu dư quý trước') or 0); xu_tong=float(p.get('Xu tổng quý') or (xu_q+xu_du)); thieu=float(p.get('Thiếu xu') or max(0,diem_q-xu_tong)); du=float(p.get('Dư xu') or max(0,xu_tong-diem_q)); tyle_q=float(p.get('Tỷ lệ quý %') or ((xu_tong/diem_q*100) if diem_q else 0))
    up_units=compare_top(a['by_unit'], b['by_unit'],5,True); down_units=compare_top(a['by_unit'], b['by_unit'],5,False); up_prod=compare_top(a['by_product'], b['by_product'],5,True); down_prod=compare_top(a['by_product'], b['by_product'],5,False)
    return f"""Kính gửi Anh/Chị {name} ({code}),

Báo cáo doanh thu Tuần 26 (01/06–26/06/2026) — phiên bản phân tích sâu.
- Doanh thu: {money(total)}; so cùng kỳ T05: {money(prev)}, chênh {money(delta)} ({pct(rate)}).
- Số dòng: {a['line_count']}; đơn vị: {a['units']}; mặt hàng: {a['products']}; nhà thầu: {a['vendors']}.
- Cơ cấu tuyến: {', '.join(f'{k}: {money(v)} ({pct(v/total*100 if total else 0)})' for k,v in sorted(a['by_tuyen'].items())) or 'chưa phát sinh'}.
- Điểm DT tháng: {num(diem_t)}; xu tháng: {num(xu_t)}.
- Điểm DT Q2: {num(diem_q)}; xu tổng Q2: {num(xu_tong)}; thiếu {num(thieu)} / dư {num(du)}; hoàn thành {pct(tyle_q)}.

Tồn tại:
- {'Doanh thu giảm '+money(abs(delta))+' so cùng kỳ T05.' if delta<0 else 'Doanh thu tăng so cùng kỳ T05, cần giữ nhịp.'}
- Đơn vị giảm/chưa phát sinh: {list_sentence(down_units,'đơn vị giảm')}.
- Mã hàng giảm/chưa phát sinh: {list_sentence(down_prod,'mã giảm')}.

Kiến nghị:
- Giữ điểm lớn: {', '.join(k for k,v in top_items(a['by_unit'],3)) or 'chưa có'}.
- Kéo lại điểm giảm: {list_sentence(down_units,'đơn vị giảm')}.
- Đẩy mã có sức kéo: {list_sentence(up_prod,'mã tăng')}.
- Chặn tụt mã cũ: {list_sentence(down_prod,'mã giảm')}.

Nguồn dữ liệu: App Report/Excel doanh số 01/06–26/06/2026 + Excel điểm/xu T06-Q2 đã đối chiếu ảnh app. Báo cáo không chứa chi phí, giá vốn, lợi nhuận, margin hoặc dữ liệu nhạy cảm.

Trân trọng,
ĐẶNG XUÂN TRUNG
CEO — Công ty TNHH Dược phẩm DONAPHARM
Hotline: 0886.396.668
"""

manifest=[]
for code in all_codes:
    h=report_html(code); t=report_txt(code)
    hp=HTML_DIR/f'week26_report_{code}.html'; tp=TXT_DIR/f'week26_report_{code}.txt'
    hp.write_text(h,encoding='utf-8'); tp.write_text(t,encoding='utf-8')
    a=aggregate(jun,code); b=aggregate(may,code); p=points.get(code,{})
    manifest.append({'code':code,'name':EMP_NAMES.get(code,code),'html':str(hp),'txt':str(tp),'revenue_jun_1_26':a['total'],'revenue_may_1_26':b['total'],'lines':a['line_count'],'units':a['units'],'products':a['products'],'vendors':a['vendors'],'point_month':p.get('Điểm tháng'),'coin_month':p.get('Xu tháng'),'point_q2':p.get('Điểm quý'),'coin_total_q2':p.get('Xu tổng quý'),'warning_lt90':p.get('Cảnh báo cá nhân <90%')})
summary={
 'generated_at':datetime.now().isoformat(timespec='seconds'),
 'source_sales_jun':str(SALES_JUN), 'source_sales_may':str(SALES_MAY), 'source_points':str(POINTS),
 'period':'01/06–26/06/2026', 'comparison':'01/05–26/05/2026',
 'employee_count':len(manifest), 'total_revenue_jun':sum(x['revenue_jun_1_26'] for x in manifest),
 'reports':manifest,
 'notes':['VP009 appears duplicated in source points file; generator keeps first occurrence.', 'No cost/gross profit/margin/CP fields used.']
}
(OUT/'manifest.json').write_text(json.dumps(summary,ensure_ascii=False,indent=2),encoding='utf-8')
print(json.dumps({'out':str(OUT),'employee_count':len(manifest),'total_revenue_jun':summary['total_revenue_jun'],'dn001_html':str(HTML_DIR/'week26_report_DN001.html'),'dn001_txt':str(TXT_DIR/'week26_report_DN001.txt')},ensure_ascii=False,indent=2))
